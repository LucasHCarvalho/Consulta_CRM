import os
import oracledb
import openpyxl
import requests
import win32com.client
from datetime import datetime

# Configurações de conexão
username = ''
password = ''
host = ''
porta = ''
service_name = ''

# Configurações de conexão para ignorar Case Sensitive na senha
oracledb.init_oracle_client()

# Crie uma conexão
conn = oracledb.connect(user=username, password=password, host=host, port=porta, service_name=service_name)
cursor = conn.cursor()

# Pesquisa Banco
cursor.execute("SELECT MEDICO FROM MEDICO")

# Retorno das pesquisas
linhas = cursor.fetchall()

# Data Atual da consulta
data_e_hora_atuais = datetime.now()
dh_atual = data_e_hora_atuais.strftime("%d/%m/%Y %H:%M")

caminho_excel = 'crm.xlsx'
arquivo_excel = openpyxl.load_workbook(caminho_excel)

# Ou use a planilha ativa padrão
planilha = arquivo_excel.active

dados_para_inserir = [
    ['Nome Cadastro SOUL', 'Nome Cadastro Cremesp', 'CRM', 'Situação SOUL', 'Situação Cremesp', 'Data Inativação Cremesp', 'Data Consulta', 'Obs']
]

for linha in linhas:
    lista_medico = [None] * 8
    lista_medico[0] = linha[0]
    lista_medico[2] = linha[1].replace(".", '')
    lista_medico[3] = linha[2]
    lista_medico[6] = dh_atual

    # URL da API com o crm a ser pesquisado
    url = f'https://api.cremesp.org.br/guia-medico/medico-info/{lista_medico[2]}'

    # Realize uma solicitação GET à API
    response = requests.get(url)

    # Verifique se a solicitação foi bem-sucedida (código de status HTTP 200)
    if response.status_code == 200:
        # Os dados da resposta da API estão em formato JSON
        dados = response.json()

        # Recebimento dos dados necessários
        lista_medico[1] = dados.get("nome")
        situacao = dados.get("situacao")
        lista_medico[5] = dados.get("mensagemStatus")
    else:
        lista_medico[1] = 'Não Encontrado'
        situacao = 'Não Encontrado'
        lista_medico[5] = 'Não Encontrado'

    if lista_medico[0].lower() != lista_medico[1].lower():
        nm_igual = False
    else:
        nm_igual = True
    
    if not nm_igual:
        lista_medico[7] = 'Nome diferente'
    
    if situacao == 'A':
        lista_medico[4] = 'ATIVO'
    else:
        lista_medico[4] = 'INATIVO'

    if lista_medico[4] == 'INATIVO' or not nm_igual:
        dados_para_inserir.append(lista_medico)

for i, dado in enumerate(dados_para_inserir, start=1):
    planilha.cell(row=i, column=1, value=dado[0])
    planilha.cell(row=i, column=2, value=dado[1])
    planilha.cell(row=i, column=3, value=dado[2])
    planilha.cell(row=i, column=4, value=dado[3])
    planilha.cell(row=i, column=5, value=dado[4])
    planilha.cell(row=i, column=6, value=dado[5])
    planilha.cell(row=i, column=7, value=dado[6])
    planilha.cell(row=i, column=8, value=dado[7])

arquivo_excel.save(caminho_excel)
arquivo_excel.close()

# Importe a classe Outlook Application
outlook = win32com.client.Dispatch("Outlook.Application")

# Crie um novo e-mail
mail = outlook.CreateItem(0)

# Anexe o arquivo CSV
anexo = os.path.abspath(caminho_excel)
mail.Attachments.Add(anexo)

# Defina o assunto do e-mail
mail.Subject = "Crm Inativo Cremesp"

# Defina o corpo do e-mail
data_e_hora_atuais = datetime.now()
data_e_hora_atuais = data_e_hora_atuais.strftime("%H:%M")

if data_e_hora_atuais < '12:00':
    saudacao = 'Bom dia!'
if data_e_hora_atuais < '18:00':
    saudacao = 'Boa tarde!'
else:
    saudacao = 'Boa noite!'

texto = ''' 

Identificamos através de uma consulta no Cremesp que os médicos da relação, encontram-se com o seu CRM inativo no Conselho Regional de Medicina.
Os mesmos estão cadastrados e com situação Ativa dentro da plataforma SOUL-MV.
'''
assinatura = '''
'''

mail.Body = saudacao + texto + assinatura

# Defina o destinatário do e-mail
mail.To = ""

# Envie o e-mail
# mail.Send()
